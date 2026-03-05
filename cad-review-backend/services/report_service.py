"""
报告生成服务模块
支持纯文字报告（plain）与图上标注报告（marked）。
"""

from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from models import Catalog, Drawing
from services.storage_path_service import resolve_project_dir

logger = logging.getLogger(__name__)


def _norm_sheet_no(value: Optional[str]) -> str:
    if not value:
        return ""
    s = str(value).strip().upper()
    s = re.sub(r"[\s\-_./\\()（）【】\[\]{}:：|]+", "", s)
    return s


def _pick_latest_drawing(rows: List[Drawing]) -> Optional[Drawing]:
    if not rows:
        return None
    return max(
        rows,
        key=lambda row: (
            row.data_version or 0,
            1 if row.png_path else 0,
            1 if row.status == "matched" else 0,
            -(row.page_index if row.page_index is not None else 10**9),
        ),
    )


def _load_evidence(result) -> Dict[str, Any]:  # noqa: ANN001
    try:
        obj = json.loads(result.evidence_json or "{}")
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass
    return {}


def _safe_float(value: Any) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _issue_code_prefix(issue_type: str) -> str:
    t = (issue_type or "").strip().lower()
    if t == "index":
        return "I"
    if t == "dimension":
        return "D"
    if t == "material":
        return "M"
    return "U"


def _severity_color(severity: str) -> Tuple[int, int, int]:
    sev = (severity or "").lower()
    if sev == "error":
        return (220, 38, 38)
    if sev == "warning":
        return (249, 115, 22)
    return (59, 130, 246)


def _issue_label(result) -> str:  # noqa: ANN001
    t = (result.type or "").strip().lower()
    if t == "dimension":
        va = (result.value_a or "").strip()
        vb = (result.value_b or "").strip()
        if va or vb:
            return f"尺寸不一致 {va or '?'} vs {vb or '?'}"
        return "尺寸不一致"
    if t == "index":
        if not (result.sheet_no_b or "").strip():
            return "索引未指向目标图"
        return "索引关系异常"
    if t == "material":
        return "材料标注不一致"
    return "发现问题"


def _shorten_text(text: str, max_chars: int = 28) -> str:
    raw = (text or "").strip()
    if len(raw) <= max_chars:
        return raw
    return raw[: max_chars - 1] + "…"


def _is_marked_pdf_enabled() -> bool:
    raw = os.getenv("REPORT_MARKED_PDF_ENABLED", "1").strip().lower()
    return raw not in {"0", "false", "no", "off"}


def _assign_issue_codes(results: List) -> List[Dict[str, Any]]:  # noqa: ANN001
    counters: Dict[str, int] = {"I": 0, "D": 0, "M": 0, "U": 0}
    rows = sorted(
        results,
        key=lambda r: (
            str(r.type or ""),
            str(r.sheet_no_a or ""),
            str(r.sheet_no_b or ""),
            str(r.location or ""),
            str(r.id or ""),
        ),
    )
    coded: List[Dict[str, Any]] = []
    for row in rows:
        prefix = _issue_code_prefix(row.type or "")
        counters[prefix] = counters.get(prefix, 0) + 1
        code = f"#{prefix}-{counters[prefix]:03d}"
        coded.append({"code": code, "result": row, "evidence": _load_evidence(row)})
    return coded


def render_annotated_png(source_png: str, marks: List[Dict[str, Any]], output_png: str) -> str:
    """
    在PNG上绘制问题标记（圆点+编号）。
    marks item: {"x": 0-100, "y": 0-100, "code": "#D-001", "severity": "warning"}
    """
    src = Path(source_png).expanduser()
    out = Path(output_png).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)

    img = Image.open(src).convert("RGB")
    draw = ImageDraw.Draw(img)
    w, h = img.size

    # 标注做“大号”，优先保证打印可读性（允许覆盖图面）
    r = max(28, int(min(w, h) * 0.016))
    label_pad_x = max(12, int(r * 0.7))
    label_pad_y = max(8, int(r * 0.5))
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Unicode.ttf", max(30, int(r * 1.5)))
    except Exception:
        font = ImageFont.load_default()

    # 同坐标问题会重叠，看起来像“没标”。先做错位展开。
    prepared: List[Dict[str, Any]] = []
    groups: Dict[Tuple[int, int], List[Dict[str, Any]]] = {}
    for mark in marks:
        x_pct = _safe_float(mark.get("x"))
        y_pct = _safe_float(mark.get("y"))
        if x_pct is None or y_pct is None:
            continue
        x = int(max(r + 4, min(w - r - 4, round(w * x_pct / 100.0))))
        y = int(max(r + 4, min(h - r - 4, round(h * y_pct / 100.0))))
        key = (int(round(x / max(1, r * 1.8))), int(round(y / max(1, r * 1.8))))
        item = dict(mark)
        item["_px"] = x
        item["_py"] = y
        groups.setdefault(key, []).append(item)

    for _, group in groups.items():
        if len(group) == 1:
            prepared.append(group[0])
            continue
        group_sorted = sorted(group, key=lambda m: str(m.get("code") or ""))
        cols = min(6, max(2, int(len(group_sorted) ** 0.5) + 1))
        step_x = int(r * 2.4)
        step_y = int(r * 2.2)
        center_col = (cols - 1) / 2.0
        base_x = int(group_sorted[0]["_px"])
        base_y = int(group_sorted[0]["_py"])
        dir_x = 1 if base_x <= int(w * 0.2) else (-1 if base_x >= int(w * 0.8) else 0)
        dir_y = 1 if base_y <= int(h * 0.2) else (-1 if base_y >= int(h * 0.8) else 0)
        for idx, item in enumerate(group_sorted):
            col = idx % cols
            row = idx // cols
            if dir_x == 0:
                ox = int((col - center_col) * step_x)
            else:
                ox = int(col * step_x * dir_x)

            if dir_y == 0:
                oy = int((row - 0.5) * step_y)
            else:
                oy = int(row * step_y * dir_y)

            item["_px"] = int(max(r + 4, min(w - r - 4, item["_px"] + ox)))
            item["_py"] = int(max(r + 4, min(h - r - 4, item["_py"] + oy)))
            prepared.append(item)

    for mark in prepared:
        x = int(mark["_px"])
        y = int(mark["_py"])

        color = _severity_color(str(mark.get("severity") or "warning"))
        code = str(mark.get("code") or "").strip() or "#U-000"
        label = _shorten_text(str(mark.get("label") or "").strip(), 30)
        text = f"{code} {label}".strip()

        draw.ellipse((x - r, y - r, x + r, y + r), outline=color, width=max(6, r // 3), fill=(255, 255, 255))
        if hasattr(draw, "textbbox"):
            bbox = draw.textbbox((0, 0), text, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        else:
            tw, th = draw.textsize(text, font=font)  # type: ignore[attr-defined]

        lx = min(max(0, x + r + 8), max(0, w - tw - label_pad_x * 2))
        ly = min(max(0, y - th // 2 - label_pad_y), max(0, h - th - label_pad_y * 2))
        draw.rectangle(
            (lx, ly, lx + tw + label_pad_x * 2, ly + th + label_pad_y * 2),
            fill=(255, 255, 255),
            outline=color,
            width=max(4, r // 4),
        )
        draw.text((lx + label_pad_x, ly + label_pad_y), text, fill=color, font=font)

    img.save(out, format="PNG")
    return str(out)


def build_sheet_issue_map(
    *,
    catalog_items: List[Catalog],
    drawing_rows: List[Drawing],
    coded_issues: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """
    返回：
    - sheets: 按目录顺序的图纸页信息（每页附 marks）
    - unlocated: 无法定位到图上的问题
    - debug_payload: 可写入 anchors.json
    """
    drawing_by_catalog: Dict[str, List[Drawing]] = {}
    drawing_by_sheet_key: Dict[str, List[Drawing]] = {}
    for row in drawing_rows:
        if row.catalog_id:
            drawing_by_catalog.setdefault(row.catalog_id, []).append(row)
        key = _norm_sheet_no(row.sheet_no)
        if key:
            drawing_by_sheet_key.setdefault(key, []).append(row)

    sheets: List[Dict[str, Any]] = []
    sheet_key_to_entry: Dict[str, Dict[str, Any]] = {}

    for item in catalog_items:
        drawing = _pick_latest_drawing(drawing_by_catalog.get(item.id, []))
        if not drawing and item.sheet_no:
            drawing = _pick_latest_drawing(drawing_by_sheet_key.get(_norm_sheet_no(item.sheet_no), []))
        entry = {
            "catalog_id": item.id,
            "sheet_no": item.sheet_no or "",
            "sheet_name": item.sheet_name or "",
            "sort_order": item.sort_order,
            "png_path": drawing.png_path if drawing else None,
            "page_index": drawing.page_index if drawing else None,
            "marks": [],
            "issue_codes": [],
        }
        sheets.append(entry)
        key = _norm_sheet_no(item.sheet_no)
        if key:
            sheet_key_to_entry[key] = entry

    # 兜底：目录外仍有图纸（极少场景）
    seen_sheet_keys = set(sheet_key_to_entry.keys())
    for row in drawing_rows:
        key = _norm_sheet_no(row.sheet_no)
        if not key or key in seen_sheet_keys:
            continue
        entry = {
            "catalog_id": None,
            "sheet_no": row.sheet_no or "",
            "sheet_name": row.sheet_name or "",
            "sort_order": 10**9,
            "png_path": row.png_path,
            "page_index": row.page_index,
            "marks": [],
            "issue_codes": [],
        }
        sheets.append(entry)
        sheet_key_to_entry[key] = entry

    unlocated: List[Dict[str, Any]] = []
    debug_issues: List[Dict[str, Any]] = []
    located_issue_count = 0

    for item in coded_issues:
        code = item["code"]
        result = item["result"]
        evidence = item["evidence"]
        anchors = evidence.get("anchors") if isinstance(evidence.get("anchors"), list) else []
        located = False
        issue_debug = {
            "code": code,
            "type": result.type,
            "severity": result.severity,
            "sheet_no_a": result.sheet_no_a,
            "sheet_no_b": result.sheet_no_b,
            "location": result.location,
            "anchors": [],
            "unlocated_reason": evidence.get("unlocated_reason"),
        }

        for anchor in anchors:
            if not isinstance(anchor, dict):
                continue
            sheet_key = _norm_sheet_no(str(anchor.get("sheet_no") or ""))
            if not sheet_key:
                continue
            sheet_entry = sheet_key_to_entry.get(sheet_key)
            if not sheet_entry:
                continue

            gp = anchor.get("global_pct") if isinstance(anchor.get("global_pct"), dict) else None
            x = _safe_float(gp.get("x")) if gp else None
            y = _safe_float(gp.get("y")) if gp else None
            if x is None or y is None:
                continue

            mark = {
                "x": x,
                "y": y,
                "code": code,
                "severity": result.severity,
                "label": _issue_label(result),
                "grid": anchor.get("grid"),
                "role": anchor.get("role"),
            }
            sheet_entry["marks"].append(mark)
            sheet_entry["issue_codes"].append(code)
            located = True
            issue_debug["anchors"].append(mark)

        if located:
            located_issue_count += 1
        else:
            unlocated.append(
                {
                    "code": code,
                    "type": result.type,
                    "severity": result.severity,
                    "sheet_no_a": result.sheet_no_a,
                    "sheet_no_b": result.sheet_no_b,
                    "location": result.location,
                    "description": result.description,
                    "reason": evidence.get("unlocated_reason") or "missing_anchor_or_global_pct",
                }
            )
        debug_issues.append(issue_debug)

    for entry in sheets:
        # 去重并保持顺序
        dedupe_codes: List[str] = []
        seen = set()
        for code in entry["issue_codes"]:
            if code in seen:
                continue
            seen.add(code)
            dedupe_codes.append(code)
        entry["issue_codes"] = dedupe_codes

    sheets.sort(key=lambda x: (x["sort_order"], x["page_index"] if x["page_index"] is not None else 10**9))

    debug_payload = {
        "generated_at": datetime.now().isoformat(),
        "issue_total": len(coded_issues),
        "located_issue_count": located_issue_count,
        "unlocated_issue_count": len(unlocated),
        "sheets": [
            {
                "sheet_no": s["sheet_no"],
                "sheet_name": s["sheet_name"],
                "png_path": s["png_path"],
                "mark_count": len(s["marks"]),
                "issue_codes": s["issue_codes"],
            }
            for s in sheets
        ],
        "issues": debug_issues,
    }
    return sheets, unlocated, debug_payload


def _build_styles():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCN", parent=styles["Title"], fontSize=22, spaceAfter=20)
    h2_style = ParagraphStyle("H2CN", parent=styles["Heading2"], fontSize=14, spaceBefore=12, spaceAfter=8)
    normal_style = ParagraphStyle("BodyCN", parent=styles["Normal"], fontSize=10, leading=14)
    return styles, title_style, h2_style, normal_style


def generate_pdf_plain(project, results: List, version: int) -> str:  # noqa: ANN001
    """
    旧版纯文字报告（兜底）。
    """
    project_dir = resolve_project_dir(project, ensure=True) / "reports"
    project_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = project_dir / f"report_v{version}.pdf"

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    _, title_style, heading_style, normal_style = _build_styles()
    story = []
    story.append(Paragraph("室内装饰施工图审核报告（文字版）", title_style))
    story.append(Paragraph(f"<b>项目名称：</b>{project.name}", normal_style))
    story.append(Paragraph(f"<b>审核日期：</b>{datetime.now().strftime('%Y年%m月%d日')}", normal_style))
    story.append(Paragraph(f"<b>审核版本：</b>V{version}", normal_style))
    story.append(Spacer(1, 8 * mm))

    index_results = [r for r in results if r.type == "index"]
    dimension_results = [r for r in results if r.type == "dimension"]
    material_results = [r for r in results if r.type == "material"]

    story.append(Paragraph(f"<b>问题总数：</b>{len(results)}", normal_style))
    story.append(Paragraph(f"<b>索引问题：</b>{len(index_results)}", normal_style))
    story.append(Paragraph(f"<b>尺寸问题：</b>{len(dimension_results)}", normal_style))
    story.append(Paragraph(f"<b>材料问题：</b>{len(material_results)}", normal_style))
    story.append(Spacer(1, 8 * mm))

    for title, rows in [("索引问题", index_results), ("尺寸问题", dimension_results), ("材料问题", material_results)]:
        story.append(Paragraph(title, heading_style))
        if not rows:
            story.append(Paragraph("✅ 未发现问题", normal_style))
        else:
            for row in rows:
                story.append(
                    Paragraph(
                        f"• [{row.severity}] {row.sheet_no_a or '-'} / {row.sheet_no_b or '-'} | {row.location or '-'} | {row.description or ''}",
                        normal_style,
                    )
                )
        story.append(Spacer(1, 4 * mm))

    doc.build(story)
    return str(pdf_path)


def generate_pdf_marked(project, results: List, version: int, db) -> Dict[str, Any]:  # noqa: ANN001
    """
    生成图上标记报告。
    """
    project_dir = resolve_project_dir(project, ensure=True) / "reports"
    project_dir.mkdir(parents=True, exist_ok=True)
    marked_pdf_path = project_dir / f"report_v{version}_marked.pdf"
    anchors_json_path = project_dir / f"report_v{version}_anchors.json"
    annotated_dir = project_dir / f"annotated_v{version}"
    annotated_dir.mkdir(parents=True, exist_ok=True)

    coded_issues = _assign_issue_codes(results)

    catalog_items = (
        db.query(Catalog)
        .filter(Catalog.project_id == project.id, Catalog.status == "locked")
        .order_by(Catalog.sort_order.asc())
        .all()
    )
    drawing_rows = (
        db.query(Drawing)
        .filter(Drawing.project_id == project.id, Drawing.replaced_at == None)  # noqa: E711
        .all()
    )

    sheets, unlocated, debug_payload = build_sheet_issue_map(
        catalog_items=catalog_items,
        drawing_rows=drawing_rows,
        coded_issues=coded_issues,
    )
    anchors_json_path.write_text(json.dumps(debug_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    located_issue_count = int(debug_payload.get("located_issue_count") or 0)
    if results and located_issue_count <= 0:
        raise RuntimeError("insufficient_anchor_evidence_for_marked_pdf")

    # 仅输出图纸页：无封面/概览/附录，避免留白和中文字体乱码。
    c = canvas.Canvas(str(marked_pdf_path), pageCompression=1)
    rendered_count = 0
    for idx, sheet in enumerate(sheets, start=1):
        png_path = sheet.get("png_path")
        if not png_path or not Path(str(png_path)).exists():
            continue

        marks = sheet.get("marks") or []
        render_path = str(png_path)
        if marks:
            safe_name = re.sub(r"[^\w\-]+", "_", f"{sheet['sheet_no']}_{idx}").strip("_") or f"sheet_{idx}"
            render_path = render_annotated_png(
                str(png_path),
                marks,
                str(annotated_dir / f"{safe_name}.png"),
            )

        with Image.open(render_path) as page_img:
            pw, ph = page_img.size
        if pw <= 0 or ph <= 0:
            continue

        # 每页尺寸和PNG一致，图面铺满无额外空白
        c.setPageSize((float(pw), float(ph)))
        c.drawImage(render_path, 0, 0, width=float(pw), height=float(ph), preserveAspectRatio=False, mask="auto")
        c.showPage()
        rendered_count += 1

    if rendered_count <= 0:
        raise RuntimeError("no_valid_png_pages_for_marked_pdf")
    c.save()

    return {
        "path": str(marked_pdf_path),
        "mode": "marked",
        "downgraded": False,
        "reason": None,
        "anchors_json_path": str(anchors_json_path),
    }


def generate_pdf(project, results: List, version: int, *, db=None, mode: str = "marked") -> Dict[str, Any]:  # noqa: ANN001
    """
    统一PDF生成入口：默认 marked，失败自动降级 plain。
    """
    mode_norm = (mode or "marked").strip().lower()
    if mode_norm not in {"marked", "plain"}:
        mode_norm = "marked"

    if mode_norm == "plain" or not _is_marked_pdf_enabled() or db is None:
        plain_path = generate_pdf_plain(project, results, version)
        reason = "mode_plain_forced" if mode_norm == "plain" else "marked_disabled_or_db_missing"
        return {
            "path": plain_path,
            "mode": "plain",
            "downgraded": mode_norm == "marked",
            "reason": reason if mode_norm == "marked" else None,
            "anchors_json_path": None,
        }

    try:
        return generate_pdf_marked(project, results, version, db)
    except Exception as exc:  # noqa: BLE001
        logger.exception("generate marked pdf failed, fallback to plain: %s", exc)
        plain_path = generate_pdf_plain(project, results, version)
        return {
            "path": plain_path,
            "mode": "plain",
            "downgraded": True,
            "reason": f"fallback_plain:{type(exc).__name__}",
            "anchors_json_path": None,
        }


def generate_excel(project, results: List, version: int) -> str:  # noqa: ANN001
    """
    生成Excel审核报告
    """
    project_dir = resolve_project_dir(project, ensure=True) / "reports"
    project_dir.mkdir(parents=True, exist_ok=True)

    excel_path = project_dir / f"report_v{version}.xlsx"
    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "概览"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")

    ws_overview.append(["项目名称", project.name])
    ws_overview.append(["审核日期", datetime.now().strftime("%Y年%m月%d日")])
    ws_overview.append(["审核版本", f"V{version}"])
    ws_overview.append([])

    index_results = [r for r in results if r.type == "index"]
    dimension_results = [r for r in results if r.type == "dimension"]
    material_results = [r for r in results if r.type == "material"]

    ws_overview.append(["问题统计"])
    ws_overview.append(["问题类型", "数量"])
    ws_overview.append(["索引问题", len(index_results)])
    ws_overview.append(["尺寸问题", len(dimension_results)])
    ws_overview.append(["材料问题", len(material_results)])
    ws_overview.append(["总计", len(results)])

    def create_sheet(wb_obj, title, data):  # noqa: ANN001
        ws = wb_obj.create_sheet(title)
        ws.append(["图号A", "图号B", "位置", "值A", "值B", "问题描述", "严重程度"])

        for cell in ws[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in data:
            severity = row.get("severity", "warning")
            fill = (
                PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                if severity == "error"
                else PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                if severity == "warning"
                else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            )

            ws.append(
                [
                    row.get("sheet_no_a", ""),
                    row.get("sheet_no_b", ""),
                    row.get("location", ""),
                    row.get("value_a", ""),
                    row.get("value_b", ""),
                    row.get("description", ""),
                    severity,
                ]
            )

            for row_cells in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
                for idx, cell in enumerate(row_cells, start=1):
                    if idx != 7:
                        cell.fill = fill
        return ws

    index_data = [
        {
            "sheet_no_a": r.sheet_no_a,
            "sheet_no_b": r.sheet_no_b,
            "location": r.location,
            "value_a": r.value_a,
            "value_b": r.value_b,
            "description": r.description,
            "severity": r.severity,
        }
        for r in index_results
    ]
    if index_data:
        create_sheet(wb, "索引问题", index_data)

    dimension_data = [
        {
            "sheet_no_a": r.sheet_no_a,
            "sheet_no_b": r.sheet_no_b,
            "location": r.location,
            "value_a": r.value_a,
            "value_b": r.value_b,
            "description": r.description,
            "severity": r.severity,
        }
        for r in dimension_results
    ]
    if dimension_data:
        create_sheet(wb, "尺寸问题", dimension_data)

    material_data = [
        {
            "sheet_no_a": r.sheet_no_a,
            "sheet_no_b": r.sheet_no_b,
            "location": r.location,
            "value_a": r.value_a,
            "value_b": r.value_b,
            "description": r.description,
            "severity": r.severity,
        }
        for r in material_results
    ]
    if material_data:
        create_sheet(wb, "材料问题", material_data)

    wb.save(str(excel_path))
    return str(excel_path)
